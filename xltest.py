import openpyxl as xl
from openpyxl.chart import BarChart, Reference


def process_worksheet(filename):
    print(xl.__version__)
    # access to file excel
    wb = xl.load_workbook('transactions2.xlsx')
    #
    # access to sheet
    sheet = wb['Sheet1']
    # access to cell have two way
    cell = sheet['a1']

    # delete colums
    # sheet.delete_cols(1, 3)
    # access all cell
    # change range 1-2 to ignore row 1(name,price,...)
    # for row in range(2, sheet.max_row + 1):
    #     # access to cell column 3
    #     cell = sheet.cell(row, 3)
    #     corrected_price = cell.value * 0.9
    #     corrected_price_cell = sheet.cell(row, 4)
    #     corrected_price_cell.value = corrected_price

    # set value in chart
    values = Reference(sheet,
                       min_row=2,
                       max_row=sheet.max_row,
                       min_col=4,
                       max_col=4)

    chart = BarChart()
    chart.add_data(values)
    # draw chart at cell e2
    sheet.add_chart(chart, 'e2')
    wb.save('transactions2.xlsx')

# wb.save('transactions2.xlsx')
